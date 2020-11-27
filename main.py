import openpyxl as xl

PATH = r'C:\python\5432192.xlsx'
NAME_SHEET = 'OTCH1'
WB = xl.load_workbook(PATH)
all_payments = {}
all_payments_keys = []


def load_values_from_xls():
    """
    fill dict all_payments by keys = date`s_of_operations, values = dict(number_of_terminal:sum_of_operation)
    data starts from row 3
    terminal number - col №8
    date of operation - col №9
    summa of operation - col №11
    """
    sheet = WB[NAME_SHEET]
    len_to_space_between_date_and_time = 11
    for i in range(3, sheet.max_row):
        terminal = sheet.cell(row=i, column=8).value
        if terminal is not None:
            terminal.strip()
        oper_date = sheet.cell(row=i, column=9).value
        if oper_date is not None:
            oper_date = oper_date[:len_to_space_between_date_and_time].strip()
        rasch_date = sheet.cell(row=i, column=10).value
        if rasch_date is not None:
            rasch_date.strip()
        summa = float(sheet.cell(row=i, column=11).value)
        if terminal:
            cur_date = all_payments.get(oper_date)
            if cur_date:
                cur_summa = cur_date.get(terminal)
                if cur_summa:
                    cur_date[terminal] = round(cur_summa + summa, 2)
                else:
                    cur_date[terminal] = round(summa, 2)
            else:
                all_payments[oper_date] = {terminal: summa}


def sort_dict_by_date():
    """
    x[3:5] - month
    x[:2] - day
    """
    global all_payments_keys
    all_payments_keys = list(all_payments.keys())
    all_payments_keys.sort(key=lambda x: (x[3:5], x[:2]))


def fill_and_save_data():
    """
    row 1 - headers
    data start from row 2
    date - col №1
    № terminal - col №2, 4, 6...
    summa by terminal - col №3, 5, 7...
    """
    if 'Результат' not in WB.sheetnames:
        WB.create_sheet(title='Результат', index=0)
        result_sheet = WB['Результат']
    else:
        result_sheet = WB['Результат']

    result_sheet.cell(row=1, column=1).value = 'Дата'
    row = 2
    col = 1
    for i in all_payments_keys:
        # print(i, ':', all_payments[i])
        result_sheet.cell(row=row, column=col).value = i

        offset = 0
        for term, term_sum in all_payments[i].items():
            result_sheet.cell(row=row, column=2+offset).value = term
            result_sheet.cell(row=1, column=2+offset).value = '№ терминала'

            result_sheet.cell(row=1, column=3+offset).value = 'Сумма по терминалу'
            result_sheet.cell(row=row, column=3+offset).value = term_sum

            offset += 2
        row += 1

    WB.save(PATH)


load_values_from_xls()
sort_dict_by_date()
fill_and_save_data()
